#!/usr/bin/env python3
"""Split each sheet of the HealthRI Excel file into a separate CSV in csv/."""

import csv
import re
from pathlib import Path

import openpyxl

EXCEL_FILE = "HealthRI Imaging metadata model v0.1.0.xlsx"
OUTPUT_DIR = Path("csv")


def sanitize_sheet_name(name: str) -> str:
    """Convert a sheet name to a safe, lowercase, underscore-separated filename."""
    slug = re.sub(r"[^a-zA-Z0-9_]", "_", name)
    slug = re.sub(r"_+", "_", slug).strip("_")
    return slug.lower()


def main():
    OUTPUT_DIR.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    written_files = set()
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        csv_name = sanitize_sheet_name(sheet_name) + ".csv"
        csv_path = OUTPUT_DIR / csv_name
        written_files.add(csv_path)

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(row)

        print(f"  {sheet_name} -> {csv_path}")

    wb.close()

    # Remove stale CSVs from deleted/renamed sheets
    for existing in OUTPUT_DIR.glob("*.csv"):
        if existing not in written_files:
            existing.unlink()
            print(f"  Removed stale: {existing}")

    print(f"Done. Wrote {len(wb.sheetnames)} CSV files to {OUTPUT_DIR}/")


if __name__ == "__main__":
    main()
